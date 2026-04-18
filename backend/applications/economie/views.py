"""Vues API pour l'économie de la construction — Plateforme LBH."""

import sys
import os
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO

from django.http import FileResponse
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from rest_framework import generics, permissions, status, filters
from rest_framework.decorators import api_view, permission_classes
from rest_framework.exceptions import ValidationError
from rest_framework.response import Response

from .models import (
    EtudeEconomique,
    LignePrix,
    EtudePrix,
    LignePrixEtude,
    AchatEtudePrix,
    ConventionCollective,
    ReferenceSocialeLocalisation,
    RegleConventionnelleProfil,
    VarianteLocaleRegleConventionnelle,
    ProfilMainOeuvre,
    AffectationProfilProjet,
)
from .serialiseurs import (
    EtudeEconomiqueListeSerialiseur,
    EtudeEconomiqueDetailSerialiseur,
    LignePrixSerialiseur,
    EtudePrixListeSerialiseur,
    EtudePrixDetailSerialiseur,
    LignePrixEtudeSerialiseur,
    AchatEtudePrixSerialiseur,
    ConventionCollectiveSerialiseur,
    ReferenceSocialeLocalisationSerialiseur,
    RegleConventionnelleProfilSerialiseur,
    VarianteLocaleRegleConventionnelleSerialiseur,
    ProfilMainOeuvreSerialiseur,
    AffectationProfilProjetSerialiseur,
    SimulationMainOeuvreEntreeSerialiseur,
    PlanActiviteEntreeSerialiseur,
)
from .services import (
    calculer_simulation_main_oeuvre,
    calculer_comparatif_estimation_etude_prix,
    construire_cadrage_etude_prix,
    generer_docx_livrable_estimation,
    generer_pdf_simulation_main_oeuvre,
    donnees_simulation_depuis_profil,
    proposer_achats_depuis_etude_prix,
    simuler_plan_activite,
)


class VueListeEtudesEconomiques(generics.ListCreateAPIView):
    """Liste et création d'études économiques."""
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["intitule", "projet__reference", "lot__intitule"]
    ordering = ["-date_modification"]

    def get_serializer_class(self):
        if self.request.method == "POST":
            return EtudeEconomiqueDetailSerialiseur
        return EtudeEconomiqueListeSerialiseur

    def get_queryset(self):
        qs = EtudeEconomique.objects.select_related("projet", "lot", "etude_parente")
        projet_id = self.request.query_params.get("projet")
        if projet_id:
            qs = qs.filter(projet_id=projet_id)
        statut = self.request.query_params.get("statut")
        if statut:
            qs = qs.filter(statut=statut)
        variantes = self.request.query_params.get("variantes")
        if variantes == "0":
            qs = qs.filter(est_variante=False)
        return qs

    def perform_create(self, serializer):
        serializer.save(cree_par=self.request.user)


class VueDetailEtudeEconomique(generics.RetrieveUpdateDestroyAPIView):
    """Détail, modification et suppression d'une étude économique."""
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = EtudeEconomiqueDetailSerialiseur

    def get_queryset(self):
        return EtudeEconomique.objects.select_related(
            "projet", "lot", "etude_parente"
        ).prefetch_related("lignes")

    def destroy(self, request, *args, **kwargs):
        etude = self.get_object()
        if etude.statut == "validee":
            return Response(
                {"detail": "Une étude validée ne peut pas être supprimée."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        etude.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class VueListeLignesPrix(generics.ListCreateAPIView):
    """Lignes de prix d'une étude économique."""
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = LignePrixSerialiseur
    ordering = ["numero_ordre"]

    def get_queryset(self):
        return LignePrix.objects.filter(
            etude_id=self.kwargs["etude_id"]
        ).select_related("ref_bibliotheque")

    def perform_create(self, serializer):
        etude = generics.get_object_or_404(EtudeEconomique, pk=self.kwargs["etude_id"])
        serializer.save(etude=etude)
        _recalculer_etude(etude)


class VueDetailLignePrix(generics.RetrieveUpdateDestroyAPIView):
    """Détail, modification et suppression d'une ligne de prix."""
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = LignePrixSerialiseur

    def get_queryset(self):
        return LignePrix.objects.filter(etude_id=self.kwargs["etude_id"])

    def perform_update(self, serializer):
        instance = serializer.save()
        _recalculer_etude(instance.etude)

    def perform_destroy(self, instance):
        etude = instance.etude
        instance.delete()
        _recalculer_etude(etude)


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
def vue_recalculer_etude(request, pk):
    """Déclenche un recalcul complet de l'étude économique."""
    etude = generics.get_object_or_404(EtudeEconomique, pk=pk)
    nb_lignes = _recalculer_etude(etude)
    return Response({
        "detail": f"Recalcul effectué : {nb_lignes} ligne(s) traitée(s).",
        "nb_lignes": nb_lignes,
    })


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
def vue_dupliquer_etude(request, pk):
    """Crée une copie de l'étude (variante ou nouvelle version)."""
    etude = generics.get_object_or_404(EtudeEconomique, pk=pk)
    est_variante = request.data.get("est_variante", False)

    nouvelle_etude = EtudeEconomique.objects.create(
        projet=etude.projet,
        lot=etude.lot,
        intitule=f"{etude.intitule} — Copie",
        statut="brouillon",
        version=etude.version + 1 if not est_variante else 1,
        est_variante=est_variante,
        etude_parente=etude if est_variante else etude.etude_parente,
        taux_frais_chantier=etude.taux_frais_chantier,
        taux_frais_generaux=etude.taux_frais_generaux,
        taux_aleas=etude.taux_aleas,
        taux_marge_cible=etude.taux_marge_cible,
        taux_pertes=etude.taux_pertes,
        cree_par=request.user,
    )

    for ligne in etude.lignes.all():
        LignePrix.objects.create(
            etude=nouvelle_etude,
            ref_bibliotheque=ligne.ref_bibliotheque,
            numero_ordre=ligne.numero_ordre,
            code=ligne.code,
            designation=ligne.designation,
            unite=ligne.unite,
            quantite_prevue=ligne.quantite_prevue,
            temps_main_oeuvre=ligne.temps_main_oeuvre,
            cout_horaire_mo=ligne.cout_horaire_mo,
            cout_matieres=ligne.cout_matieres,
            cout_materiel=ligne.cout_materiel,
            cout_sous_traitance=ligne.cout_sous_traitance,
            cout_transport=ligne.cout_transport,
            taux_pertes_surcharge=ligne.taux_pertes_surcharge,
            taux_frais_chantier_surcharge=ligne.taux_frais_chantier_surcharge,
            taux_frais_generaux_surcharge=ligne.taux_frais_generaux_surcharge,
            taux_aleas_surcharge=ligne.taux_aleas_surcharge,
            taux_marge_surcharge=ligne.taux_marge_surcharge,
            observations=ligne.observations,
        )

    _recalculer_etude(nouvelle_etude)

    serialiseur = EtudeEconomiqueDetailSerialiseur(
        nouvelle_etude, context={"request": request}
    )
    return Response(serialiseur.data, status=status.HTTP_201_CREATED)


def _recalculer_etude(etude: EtudeEconomique) -> int:
    """
    Appelle le moteur de rentabilité sur toutes les lignes de l'étude
    et met à jour les totaux. Retourne le nombre de lignes traitées.
    """
    try:
        # Chemin vers le moteur de calcul (hors du package Django)
        racine = os.path.dirname(
            os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        )
        if racine not in sys.path:
            sys.path.insert(0, racine)
        from calculs.economie.moteur_rentabilite import (
            calculer_ligne,
            ParametresCalcul,
            ComposantesDebourse,
        )
        from applications.parametres.models import Parametre

        def _taux(code: str, defaut: str) -> Decimal:
            try:
                return Decimal(str(Parametre.objects.get(cle=code).valeur_typee()))
            except Exception:
                return Decimal(defaut)

        params = ParametresCalcul(
            taux_frais_chantier=Decimal(str(etude.taux_frais_chantier)) if etude.taux_frais_chantier else _taux("TAUX_FRAIS_CHANTIER", "0.08"),
            taux_frais_generaux=Decimal(str(etude.taux_frais_generaux)) if etude.taux_frais_generaux else _taux("TAUX_FRAIS_GENERAUX", "0.12"),
            taux_aleas=Decimal(str(etude.taux_aleas)) if etude.taux_aleas else _taux("TAUX_ALEAS", "0.03"),
            taux_marge_cible=Decimal(str(etude.taux_marge_cible)) if etude.taux_marge_cible else _taux("TAUX_MARGE_CIBLE", "0.10"),
            taux_pertes=Decimal(str(etude.taux_pertes)) if etude.taux_pertes else _taux("TAUX_PERTES_MATIERES", "0.05"),
        )

        lignes = list(etude.lignes.all())
        if not lignes:
            return 0

        # Calcul de chaque ligne via le moteur
        resultats = []
        for ligne in lignes:
            if not ligne.quantite_prevue or ligne.quantite_prevue <= 0:
                continue
            composantes = ComposantesDebourse(
                temps_main_oeuvre=ligne.temps_main_oeuvre,
                cout_horaire_mo=ligne.cout_horaire_mo,
                cout_matieres=ligne.cout_matieres,
                cout_materiel=ligne.cout_materiel,
                cout_sous_traitance=ligne.cout_sous_traitance,
                cout_transport=ligne.cout_transport,
            )
            res = calculer_ligne(
                composantes=composantes,
                quantite_prevue=ligne.quantite_prevue,
                params=params,
                taux_pertes_surcharge=ligne.taux_pertes_surcharge,
                taux_frais_chantier_surcharge=ligne.taux_frais_chantier_surcharge,
                taux_frais_generaux_surcharge=ligne.taux_frais_generaux_surcharge,
                taux_aleas_surcharge=ligne.taux_aleas_surcharge,
                taux_marge_surcharge=ligne.taux_marge_surcharge,
                quantite_reelle=ligne.quantite_reelle,
            )
            resultats.append((ligne, res))

        # Calcul des totaux pour contribution_marge
        total_marge_nette = sum(r.marge_nette_totale for _, r in resultats)
        total_prix_vente = sum(r.prix_vente_unitaire * r.quantite_prevue for _, r in resultats)

        # Application des résultats sur les lignes
        a_sauvegarder = []
        for ligne, res in resultats:
            ligne.debourse_sec_unitaire = res.debourse_sec_unitaire
            ligne.cout_direct_unitaire = res.cout_direct_unitaire
            ligne.cout_revient_unitaire = res.cout_revient_unitaire
            ligne.prix_vente_unitaire = res.prix_vente_unitaire
            ligne.marge_brute_unitaire = res.marge_brute_unitaire
            ligne.marge_nette_unitaire = res.marge_nette_unitaire
            ligne.taux_marge_nette = res.taux_marge_nette
            ligne.marge_brute_totale = res.marge_brute_totale
            ligne.marge_nette_totale = res.marge_nette_totale
            ligne.etat_rentabilite = res.etat_rentabilite.value
            ligne.seuil_quantite_critique = res.seuil_quantite_critique
            ligne.seuil_prix_minimum = res.seuil_prix_minimum
            ligne.causes_non_rentabilite = res.causes_non_rentabilite
            ligne.indice_sensibilite_quantite = res.indice_sensibilite_quantite
            ligne.indice_sensibilite_main_oeuvre = res.indice_sensibilite_main_oeuvre
            ligne.indice_sensibilite_matieres = res.indice_sensibilite_matieres
            if total_marge_nette and total_marge_nette != 0:
                ligne.contribution_marge = (res.marge_nette_totale / total_marge_nette).quantize(
                    Decimal("0.0001"), rounding=ROUND_HALF_UP,
                )
            else:
                ligne.contribution_marge = Decimal("0")
            a_sauvegarder.append(ligne)

        # Mise à jour en masse
        if a_sauvegarder:
            LignePrix.objects.bulk_update(a_sauvegarder, [
                "debourse_sec_unitaire", "cout_direct_unitaire",
                "cout_revient_unitaire", "prix_vente_unitaire",
                "marge_brute_unitaire", "marge_nette_unitaire", "taux_marge_nette",
                "marge_brute_totale", "marge_nette_totale", "contribution_marge",
                "etat_rentabilite", "seuil_quantite_critique", "seuil_prix_minimum",
                "causes_non_rentabilite",
                "indice_sensibilite_quantite", "indice_sensibilite_main_oeuvre",
                "indice_sensibilite_matieres",
            ])

        # Agrégation des totaux pour l'étude
        total_ds = sum(r.debourse_sec_total for _, r in resultats)
        total_cd = sum(r.cout_direct_unitaire * r.quantite_prevue for _, r in resultats)
        total_cr = sum(r.cout_revient_unitaire * r.quantite_prevue for _, r in resultats)
        total_mb = sum(r.marge_brute_totale for _, r in resultats)
        tau_global = (
            (total_marge_nette / total_prix_vente).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)
            if total_prix_vente else Decimal("0")
        )

        etude.total_debourse_sec = total_ds
        etude.total_cout_direct = total_cd
        etude.total_cout_revient = total_cr
        etude.total_prix_vente = total_prix_vente
        etude.total_marge_brute = total_mb
        etude.total_marge_nette = total_marge_nette
        etude.taux_marge_nette_global = tau_global
        etude.save(update_fields=[
            "total_debourse_sec", "total_cout_direct", "total_cout_revient",
            "total_prix_vente", "total_marge_brute", "total_marge_nette",
            "taux_marge_nette_global",
        ])
        return len(a_sauvegarder)

    except ImportError:
        # Moteur non encore disponible — calcul différé
        return 0


# ─── Études de prix ───────────────────────────────────────────────────────────

class VueListeEtudesPrix(generics.ListCreateAPIView):
    """Liste et création d'études de prix."""
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["code", "intitule", "lot_type"]
    ordering = ["-date_modification"]

    def get_serializer_class(self):
        if self.request.method == "POST":
            return EtudePrixDetailSerialiseur
        return EtudePrixListeSerialiseur

    def get_queryset(self):
        qs = EtudePrix.objects.select_related("projet", "organisation", "auteur")
        projet_id = self.request.query_params.get("projet")
        if projet_id:
            qs = qs.filter(projet_id=projet_id)
        statut = self.request.query_params.get("statut")
        if statut:
            qs = qs.filter(statut=statut)
        lot_type = self.request.query_params.get("lot_type")
        if lot_type:
            qs = qs.filter(lot_type=lot_type)
        millesime = self.request.query_params.get("millesime")
        if millesime:
            qs = qs.filter(millesime=millesime)
        return qs

    def perform_create(self, serializer):
        serializer.save(auteur=self.request.user)


class VueDetailEtudePrix(generics.RetrieveUpdateDestroyAPIView):
    """Détail, modification et archivage d'une étude de prix."""
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = EtudePrixDetailSerialiseur

    def get_queryset(self):
        return EtudePrix.objects.select_related(
            "projet", "organisation", "auteur", "validateur", "ligne_bibliotheque"
        ).prefetch_related("lignes__profil_main_oeuvre", "achats__ligne_source")

    def perform_update(self, serializer):
        etude = self.get_object()
        _verifier_etude_prix_modifiable(etude)
        etude = serializer.save()
        etude.recalculer_totaux()

    def destroy(self, request, *args, **kwargs):
        etude = self.get_object()
        if etude.statut in ("publiee", "validee"):
            return Response(
                {"detail": "Une étude validée ou publiée ne peut pas être supprimée."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        etude.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class VueListeLignesPrixEtude(generics.ListCreateAPIView):
    """Lignes de ressource d'une étude de prix."""
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = LignePrixEtudeSerialiseur
    ordering = ["ordre"]

    def get_queryset(self):
        return LignePrixEtude.objects.select_related("profil_main_oeuvre").filter(etude_id=self.kwargs["etude_id"])

    def perform_create(self, serializer):
        etude = generics.get_object_or_404(EtudePrix, pk=self.kwargs["etude_id"])
        _verifier_etude_prix_modifiable(etude)
        serializer.save(etude=etude)
        etude.recalculer_totaux()
        _mettre_etude_prix_en_cours_si_necessaire(etude)


class VueDetailLignePrixEtude(generics.RetrieveUpdateDestroyAPIView):
    """Détail, modification et suppression d'une ligne de ressource."""
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = LignePrixEtudeSerialiseur

    def get_queryset(self):
        return LignePrixEtude.objects.select_related("profil_main_oeuvre").filter(etude_id=self.kwargs["etude_id"])

    def perform_update(self, serializer):
        _verifier_etude_prix_modifiable(serializer.instance.etude)
        instance = serializer.save()
        instance.etude.recalculer_totaux()
        _mettre_etude_prix_en_cours_si_necessaire(instance.etude)

    def perform_destroy(self, instance):
        etude = instance.etude
        _verifier_etude_prix_modifiable(etude)
        instance.delete()
        etude.recalculer_totaux()


class VueListeAchatsEtudePrix(generics.ListCreateAPIView):
    """Achats fournisseurs et conditionnements d'une étude de prix."""
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = AchatEtudePrixSerialiseur
    ordering = ["ordre", "designation"]

    def get_queryset(self):
        return AchatEtudePrix.objects.select_related("ligne_source").filter(etude_id=self.kwargs["etude_id"])

    def perform_create(self, serializer):
        etude = generics.get_object_or_404(EtudePrix, pk=self.kwargs["etude_id"])
        _verifier_etude_prix_modifiable(etude)
        serializer.save(etude=etude)


class VueDetailAchatEtudePrix(generics.RetrieveUpdateDestroyAPIView):
    """Détail d'un achat fournisseur rattaché à une étude de prix."""
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = AchatEtudePrixSerialiseur

    def get_queryset(self):
        return AchatEtudePrix.objects.select_related("ligne_source").filter(etude_id=self.kwargs["etude_id"])

    def perform_update(self, serializer):
        _verifier_etude_prix_modifiable(serializer.instance.etude)
        serializer.save()

    def perform_destroy(self, instance):
        _verifier_etude_prix_modifiable(instance.etude)
        instance.delete()


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
def vue_valider_etude_prix(request, pk):
    """Valide une étude de prix (passage en statut 'validée')."""
    etude = generics.get_object_or_404(EtudePrix, pk=pk)
    if etude.statut == "brouillon" and etude.lignes.exists():
        etude.statut = "en_cours"
        etude.save(update_fields=["statut"])
    if etude.statut not in ("en_cours", "a_valider"):
        return Response(
            {"detail": "Seule une étude en cours ou à valider peut être validée."},
            status=status.HTTP_400_BAD_REQUEST,
        )
    etude.statut = "validee"
    etude.validateur = request.user
    from django.utils import timezone
    etude.date_validation = timezone.now().date()
    etude.save(update_fields=["statut", "validateur", "date_validation"])
    return Response({"detail": "Étude de prix validée."})


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
def vue_publier_etude_prix(request, pk):
    """
    Publie une étude de prix validée vers la bibliothèque de prix.
    Crée ou met à jour une LignePrixBibliotheque depuis les résultats de l'étude.
    """
    from applications.bibliotheque.models import LignePrixBibliotheque
    from applications.bibliotheque.services import (
        recalculer_composantes_depuis_sous_details,
        synchroniser_article_cctp_reference,
        synchroniser_sous_details_depuis_etude_prix,
    )

    etude = generics.get_object_or_404(EtudePrix, pk=pk)
    if etude.statut != "validee":
        return Response(
            {"detail": "Seule une étude validée peut être publiée en bibliothèque."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    etude.recalculer_totaux()

    # Création ou mise à jour de la ligne bibliothèque
    if etude.ligne_bibliotheque:
        entree = etude.ligne_bibliotheque
    else:
        entree = LignePrixBibliotheque(
            niveau="reference",
            famille=etude.get_lot_type_display() or "Général",
            lot=etude.get_lot_type_display() or "",
        )

    entree.code = etude.code or f"EP-{str(etude.id)[:8].upper()}"
    entree.designation_courte = etude.intitule[:300]
    entree.designation_longue = etude.description or etude.intitule
    entree.unite = request.data.get("unite", "U")
    entree.hypotheses = etude.hypotheses
    entree.observations_economiques = etude.observations

    # Ventilation des coûts
    entree.cout_horaire_mo = etude.taux_horaire_mo
    # Coût MO = total_mo en temps (si la quantité de l'ouvrage est fournie)
    quantite_ouvrage = Decimal(str(request.data.get("quantite_ouvrage", 1)))
    if quantite_ouvrage > 0:
        entree.cout_matieres = (etude.total_matieres_ht / quantite_ouvrage).quantize(Decimal("0.0001"))
        entree.cout_materiel = (etude.total_materiel_ht / quantite_ouvrage).quantize(Decimal("0.0001"))
        entree.cout_sous_traitance = (etude.total_sous_traitance_ht / quantite_ouvrage).quantize(Decimal("0.0001"))
        entree.cout_transport = (etude.total_transport_ht / quantite_ouvrage).quantize(Decimal("0.0001"))
        entree.cout_frais_divers = (etude.total_frais_divers_ht / quantite_ouvrage).quantize(Decimal("0.0001"))
        entree.debourse_sec_unitaire = (etude.debourse_sec_ht / quantite_ouvrage).quantize(Decimal("0.0001"))

        # Temps MO unitaire depuis les lignes MO
        total_heures_mo = sum(
            l.quantite for l in etude.lignes.filter(type_ressource="mo")
        )
        if quantite_ouvrage > 0:
            entree.temps_main_oeuvre = (total_heures_mo / quantite_ouvrage).quantize(Decimal("0.0001"))

    entree.source = f"Étude de prix — {etude.intitule} ({etude.millesime})"
    entree.origine_import = "etude_prix"
    entree.fiabilite = 4  # Issu d'étude analytique
    entree.statut_validation = "valide"
    entree.save()

    synchroniser_sous_details_depuis_etude_prix(etude, entree, quantite_ouvrage=quantite_ouvrage)
    composantes = recalculer_composantes_depuis_sous_details(entree)
    for champ, valeur in composantes.items():
        setattr(entree, champ, valeur)
    entree.save(update_fields=list(composantes.keys()) + ["origine_import", "date_modification"])
    synchroniser_article_cctp_reference(entree)

    etude.ligne_bibliotheque = entree
    etude.statut = "publiee"
    etude.save(update_fields=["ligne_bibliotheque", "statut"])

    return Response({
        "detail": "Étude publiée en bibliothèque.",
        "ligne_bibliotheque_id": str(entree.id),
    })


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated])
def vue_exporter_etude_prix_xlsx(request, pk):
    """Exporte l'étude de prix sous forme de tableur XLSX exploitable comme base DPGF/BPU."""
    etude = generics.get_object_or_404(
        EtudePrix.objects.select_related("projet", "organisation", "auteur", "ligne_bibliotheque").prefetch_related("lignes"),
        pk=pk,
    )
    etude.recalculer_totaux()

    classeur = Workbook()
    feuille = classeur.active
    feuille.title = "Bordereau"
    feuille.append(["Code", "Désignation", "Type", "Unité", "Quantité", "Coût unitaire HT", "Montant HT", "Observations"])

    for index, ligne in enumerate(etude.lignes.order_by("ordre"), start=2):
        feuille.cell(row=index, column=1, value=ligne.code)
        feuille.cell(row=index, column=2, value=ligne.designation)
        feuille.cell(row=index, column=3, value=ligne.get_type_ressource_display())
        feuille.cell(row=index, column=4, value=ligne.unite)
        feuille.cell(row=index, column=5, value=float(ligne.quantite))
        feuille.cell(row=index, column=6, value=float(ligne.cout_unitaire_ht))
        feuille.cell(row=index, column=7, value=f"=E{index}*F{index}")
        feuille.cell(row=index, column=8, value=ligne.observations)

    total_row = feuille.max_row + 2
    feuille.cell(row=total_row, column=6, value="Déboursé sec HT")
    feuille.cell(row=total_row, column=7, value=f"=SUM(G2:G{total_row - 2})")

    synthese = classeur.create_sheet("Synthèse")
    synthese.append(["Champ", "Valeur"])
    synthese.append(["Intitulé", etude.intitule])
    synthese.append(["Code", etude.code])
    synthese.append(["Projet", etude.projet.reference if etude.projet else ""])
    synthese.append(["Méthode", etude.get_methode_display()])
    synthese.append(["Lot", etude.get_lot_type_display() if etude.lot_type else ""])
    synthese.append(["Millésime", etude.millesime])
    synthese.append(["Taux horaire MO", float(etude.taux_horaire_mo)])
    synthese.append(["Hypothèses", etude.hypotheses])
    synthese.append(["Observations", etude.observations])
    synthese.append(["Total MO HT", float(etude.total_mo_ht)])
    synthese.append(["Total matières HT", float(etude.total_matieres_ht)])
    synthese.append(["Total matériel HT", float(etude.total_materiel_ht)])
    synthese.append(["Total sous-traitance HT", float(etude.total_sous_traitance_ht)])
    synthese.append(["Total transport HT", float(etude.total_transport_ht)])
    synthese.append(["Total frais divers HT", float(etude.total_frais_divers_ht)])
    synthese.append(["Déboursé sec HT", float(etude.debourse_sec_ht)])
    synthese.append(["Frais de chantier HT", float(etude.montant_frais_chantier_ht)])
    synthese.append(["Frais généraux HT", float(etude.montant_frais_generaux_ht)])
    synthese.append(["Aléas HT", float(etude.montant_aleas_ht)])
    synthese.append(["Coût de revient HT", float(etude.cout_revient_ht)])
    synthese.append(["Marge prévisionnelle HT", float(etude.marge_previsionnelle_ht)])
    synthese.append(["Prix de vente HT", float(etude.prix_vente_ht)])
    synthese.append(["Coefficient K", float(etude.coefficient_k)])
    synthese.append(["Seuil de rentabilité HT", float(etude.seuil_rentabilite_ht)])

    sortie = BytesIO()
    classeur.save(sortie)
    sortie.seek(0)

    fichier = f"{(etude.code or 'etude-prix').replace('/', '-')}.xlsx"
    return FileResponse(
        sortie,
        as_attachment=True,
        filename=fichier,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated])
def vue_comparatif_etude_prix(request, pk):
    """Retourne les estimations ratio/REX/analytique et les projets similaires."""
    etude = generics.get_object_or_404(
        EtudePrix.objects.select_related("projet", "organisation", "auteur").prefetch_related("lignes", "achats"),
        pk=pk,
    )
    etude.recalculer_totaux()
    return Response(calculer_comparatif_estimation_etude_prix(etude))


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated])
def vue_exporter_comparatif_etude_prix_xlsx(request, pk):
    """Exporte la comparaison ratio / retour d'expérience / analytique avec graphique."""
    etude = generics.get_object_or_404(
        EtudePrix.objects.select_related("projet", "organisation", "auteur").prefetch_related("lignes", "achats"),
        pk=pk,
    )
    etude.recalculer_totaux()
    comparatif = calculer_comparatif_estimation_etude_prix(etude)

    classeur = Workbook()
    feuille = classeur.active
    feuille.title = "Comparatif"
    feuille.append(["Indicateur", "Valeur HT"])
    feuille.append(["Déboursé sec", float(etude.debourse_sec_ht)])
    feuille.append(["Seuil de rentabilité", float(etude.seuil_rentabilite_ht)])
    feuille.append(["Estimation ratio", float(comparatif["estimation_ratio_ht"])])
    feuille.append(["Estimation retour d'expérience", float(comparatif["estimation_rex_ht"])])
    feuille.append(["Estimation analytique", float(comparatif["estimation_analytique_ht"])])
    feuille.append(["Prix de vente prévisionnel", float(etude.prix_vente_ht)])

    graphique = BarChart()
    graphique.type = "col"
    graphique.style = 10
    graphique.title = "Comparaison des estimations"
    graphique.y_axis.title = "Montant HT"
    graphique.x_axis.title = "Indicateurs"
    data = Reference(feuille, min_col=2, min_row=1, max_row=7)
    categories = Reference(feuille, min_col=1, min_row=2, max_row=7)
    graphique.add_data(data, titles_from_data=True)
    graphique.set_categories(categories)
    graphique.height = 8
    graphique.width = 16
    feuille.add_chart(graphique, "D2")

    feuille_similaires = classeur.create_sheet("Projets similaires")
    feuille_similaires.append([
        "Référence", "Projet", "Type", "Clientèle", "Phase", "Département",
        "Surface", "Montant", "Score similarité",
    ])
    for projet in comparatif["projets_similaires"]:
        clientele = projet.get("clientele_cible", "")
        feuille_similaires.append([
            projet.get("reference", ""),
            projet.get("intitule", ""),
            projet.get("type_projet", ""),
            clientele,
            projet.get("phase_actuelle", ""),
            projet.get("departement", ""),
            float(projet.get("surface_reference") or 0),
            float(projet.get("montant_reference") or 0),
            float(projet.get("score_similarite") or 0),
        ])

    feuille_alertes = classeur.create_sheet("Alertes")
    feuille_alertes.append(["Message"])
    for alerte in comparatif["alertes"] or ["Aucune alerte."]:
        feuille_alertes.append([alerte])

    sortie = BytesIO()
    classeur.save(sortie)
    sortie.seek(0)
    fichier = f"{(etude.code or 'comparatif-etude-prix').replace('/', '-')}-comparatif.xlsx"
    return FileResponse(
        sortie,
        as_attachment=True,
        filename=fichier,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated])
def vue_exporter_achats_etude_prix_xlsx(request, pk):
    """Exporte un bon de commande fournisseurs synthétique en XLSX."""
    etude = generics.get_object_or_404(
        EtudePrix.objects.select_related("projet", "organisation").prefetch_related("achats__ligne_source"),
        pk=pk,
    )
    classeur = Workbook()
    feuille = classeur.active
    feuille.title = "Bon de commande"
    feuille.append([
        "Ordre", "Désignation", "Fournisseur", "Référence fournisseur", "Ligne source",
        "Unité", "Besoin", "Conditionnement", "Nb conditionnements", "Qté commandée",
        "Prix unitaire achat HT", "Coût total achat HT", "Surcoût conditionnement HT", "Observations",
    ])

    total = Decimal("0")
    for achat in etude.achats.order_by("ordre", "designation"):
        total += Decimal(str(achat.cout_total_achat_ht or 0))
        feuille.append([
            achat.ordre,
            achat.designation,
            achat.fournisseur,
            achat.reference_fournisseur,
            achat.ligne_source.designation if achat.ligne_source else "",
            achat.unite_achat,
            float(achat.quantite_besoin),
            float(achat.quantite_conditionnement),
            float(achat.nombre_conditionnements),
            float(achat.quantite_commandee),
            float(achat.prix_unitaire_achat_ht),
            float(achat.cout_total_achat_ht),
            float(achat.surcout_conditionnement_ht),
            achat.observations,
        ])

    feuille.append([])
    feuille.append(["", "", "", "", "", "", "", "", "", "", "Total achats HT", float(total)])

    sortie = BytesIO()
    classeur.save(sortie)
    sortie.seek(0)
    fichier = f"{(etude.code or 'etude-prix').replace('/', '-')}-bon-commande.xlsx"
    return FileResponse(
        sortie,
        as_attachment=True,
        filename=fichier,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _exporter_note_estimation_docx(pk, *, mode: str):
    etude = generics.get_object_or_404(
        EtudePrix.objects.select_related("projet", "organisation", "auteur").prefetch_related("lignes", "achats"),
        pk=pk,
    )
    etude.recalculer_totaux()
    comparatif = calculer_comparatif_estimation_etude_prix(etude)
    contenu = generer_docx_livrable_estimation(etude, comparatif, mode=mode)
    suffixe = "note-moa" if mode == "moa" else "note-moe"
    fichier = f"{(etude.code or 'etude-prix').replace('/', '-')}-{suffixe}.docx"
    return FileResponse(
        BytesIO(contenu),
        as_attachment=True,
        filename=fichier,
        content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    )


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated])
def vue_exporter_note_moa_etude_prix_docx(request, pk):
    """Exporte une note de vérification d'enveloppe orientée MOA."""
    return _exporter_note_estimation_docx(pk, mode="moa")


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated])
def vue_exporter_note_moe_etude_prix_docx(request, pk):
    """Exporte une note d'estimation consolidée orientée MOE."""
    return _exporter_note_estimation_docx(pk, mode="moe")


def _mettre_etude_prix_en_cours_si_necessaire(etude: EtudePrix) -> None:
    """
    Une étude de prix qui contient des ressources ne doit pas rester en brouillon.
    Cela rétablit un cycle cohérent entre saisie, validation et publication.
    """
    if etude.statut == "brouillon" and etude.lignes.exists():
        etude.statut = "en_cours"
        etude.save(update_fields=["statut"])


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
def vue_proposer_achats_etude_prix(request, pk):
    """Crée ou remplace les achats proposés à partir des ressources matières."""
    etude = generics.get_object_or_404(EtudePrix, pk=pk)
    _verifier_etude_prix_modifiable(etude)
    remplacer = str(request.data.get("remplacer", "")).lower() in {"1", "true", "oui", "yes"}
    achats = proposer_achats_depuis_etude_prix(etude, remplacer=remplacer)
    return Response({
        "detail": f"{len(achats)} achat(s) proposé(s) depuis les fournitures de l'étude.",
        "achats": AchatEtudePrixSerialiseur(achats, many=True).data,
    })


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated])
def vue_cadrage_etude_prix(request, pk):
    """Retourne le cadrage métier et les questions guidées d'une étude de prix."""
    etude = generics.get_object_or_404(
        EtudePrix.objects.select_related("projet", "organisation").prefetch_related("lignes", "achats"),
        pk=pk,
    )
    return Response(construire_cadrage_etude_prix(etude))


def _verifier_etude_prix_modifiable(etude: EtudePrix) -> None:
    """Empêche la modification des études de prix finalisées."""
    if etude.statut in ("validee", "publiee", "archivee"):
        raise ValidationError(
            "Cette étude de prix est finalisée et ne peut plus être modifiée. "
            "Dupliquez-la pour créer une nouvelle variante ou version."
        )


# ─── Profils de main-d'œuvre et simulateur ──────────────────────────────────


class VueListeConventionsCollectives(generics.ListCreateAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = ConventionCollectiveSerialiseur
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["code", "libelle", "idcc"]
    ordering = ["libelle"]

    def get_queryset(self):
        qs = ConventionCollective.objects.all()
        actifs = self.request.query_params.get("actifs")
        if actifs == "1":
            qs = qs.filter(est_active=True)
        return qs


class VueDetailConventionCollective(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = ConventionCollectiveSerialiseur
    queryset = ConventionCollective.objects.all()


class VueListeReferencesSocialesLocalisation(generics.ListCreateAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = ReferenceSocialeLocalisationSerialiseur
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["code", "libelle", "localisation"]
    ordering = ["localisation", "libelle"]

    def get_queryset(self):
        qs = ReferenceSocialeLocalisation.objects.all()
        actifs = self.request.query_params.get("actifs")
        if actifs == "1":
            qs = qs.filter(est_active=True)
        return qs


class VueDetailReferenceSocialeLocalisation(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = ReferenceSocialeLocalisationSerialiseur
    queryset = ReferenceSocialeLocalisation.objects.all()


class VueListeReglesConventionnelles(generics.ListCreateAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = RegleConventionnelleProfilSerialiseur
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["code", "libelle", "niveau_classification"]
    ordering = ["convention__libelle", "ordre_affichage", "libelle"]

    def get_queryset(self):
        qs = RegleConventionnelleProfil.objects.select_related("convention").prefetch_related("variantes_locales")
        convention_id = self.request.query_params.get("convention")
        if convention_id:
            qs = qs.filter(convention_id=convention_id)
        actifs = self.request.query_params.get("actifs")
        if actifs == "1":
            qs = qs.filter(est_active=True, convention__est_active=True)
        return qs


class VueDetailRegleConventionnelle(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = RegleConventionnelleProfilSerialiseur
    queryset = RegleConventionnelleProfil.objects.select_related("convention").prefetch_related("variantes_locales")


class VueListeVariantesLocalesReglesConventionnelles(generics.ListCreateAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = VarianteLocaleRegleConventionnelleSerialiseur
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["libelle", "regle__libelle", "localisation"]
    ordering = ["regle__libelle", "localisation"]

    def get_queryset(self):
        qs = VarianteLocaleRegleConventionnelle.objects.select_related("regle", "regle__convention")
        regle_id = self.request.query_params.get("regle")
        if regle_id:
            qs = qs.filter(regle_id=regle_id)
        localisation = self.request.query_params.get("localisation")
        if localisation:
            qs = qs.filter(localisation=localisation)
        actifs = self.request.query_params.get("actifs")
        if actifs == "1":
            qs = qs.filter(est_active=True)
        return qs


class VueDetailVarianteLocaleRegleConventionnelle(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = VarianteLocaleRegleConventionnelleSerialiseur
    queryset = VarianteLocaleRegleConventionnelle.objects.select_related("regle", "regle__convention")


class VueListeProfilsMainOeuvre(generics.ListCreateAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = ProfilMainOeuvreSerialiseur
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["code", "libelle", "categorie"]
    ordering = ["ordre_affichage", "libelle"]

    def get_queryset(self):
        qs = ProfilMainOeuvre.objects.select_related("convention_collective", "regle_conventionnelle")
        actifs = self.request.query_params.get("actifs")
        if actifs == "1":
            qs = qs.filter(est_actif=True)
        secteur_activite = self.request.query_params.get("secteur_activite")
        if secteur_activite:
            qs = qs.filter(secteur_activite=secteur_activite)
        return qs


class VueDetailProfilMainOeuvre(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = ProfilMainOeuvreSerialiseur
    queryset = ProfilMainOeuvre.objects.select_related("convention_collective", "regle_conventionnelle")


class VueListeAffectationsProfilsProjet(generics.ListCreateAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = AffectationProfilProjetSerialiseur
    ordering = ["-date_modification"]

    def get_queryset(self):
        qs = AffectationProfilProjet.objects.select_related("projet", "profil")
        projet_id = self.request.query_params.get("projet")
        if projet_id:
            qs = qs.filter(projet_id=projet_id)
        return qs


class VueDetailAffectationProfilProjet(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = AffectationProfilProjetSerialiseur
    queryset = AffectationProfilProjet.objects.select_related("projet", "profil")


def _donnees_simulation_hydratees(donnees_validees: dict) -> dict:
    profil_code = donnees_validees.get("profil_code")
    if not profil_code:
        return donnees_validees

    try:
        profil = ProfilMainOeuvre.objects.select_related(
            "convention_collective",
            "regle_conventionnelle",
        ).prefetch_related(
            "regle_conventionnelle__variantes_locales",
        ).get(code=profil_code)
    except ProfilMainOeuvre.DoesNotExist:
        return donnees_validees

    return donnees_simulation_depuis_profil(profil, donnees_validees)


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated])
def vue_recuperer_defauts_simulation_profil(request, pk):
    profil = generics.get_object_or_404(
        ProfilMainOeuvre.objects.select_related("convention_collective", "regle_conventionnelle").prefetch_related(
            "regle_conventionnelle__variantes_locales"
        ),
        pk=pk,
    )
    localisation = request.query_params.get("localisation") or profil.localisation
    return Response(donnees_simulation_depuis_profil(profil, {"localisation": localisation}))


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
def vue_simuler_cout_main_oeuvre(request):
    serialiseur = SimulationMainOeuvreEntreeSerialiseur(data=request.data)
    serialiseur.is_valid(raise_exception=True)
    return Response(calculer_simulation_main_oeuvre(_donnees_simulation_hydratees(serialiseur.validated_data)))


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
def vue_exporter_simulation_main_oeuvre_pdf(request):
    serialiseur = SimulationMainOeuvreEntreeSerialiseur(data=request.data)
    serialiseur.is_valid(raise_exception=True)
    simulation = calculer_simulation_main_oeuvre(_donnees_simulation_hydratees(serialiseur.validated_data))
    contenu = generer_pdf_simulation_main_oeuvre(simulation)
    reponse = FileResponse(
        BytesIO(contenu),
        as_attachment=True,
        filename="fiche-simulation-main-oeuvre.pdf",
        content_type="application/pdf",
    )
    reponse["Content-Length"] = str(len(contenu))
    return reponse


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
def vue_creer_affectation_depuis_simulation(request):
    projet_id = request.data.get("projet")
    profil_id = request.data.get("profil")
    if not projet_id or not profil_id:
        return Response(
            {"detail": "Les champs « projet » et « profil » sont requis."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    serialiseur = SimulationMainOeuvreEntreeSerialiseur(data=request.data)
    serialiseur.is_valid(raise_exception=True)
    simulation = calculer_simulation_main_oeuvre(_donnees_simulation_hydratees(serialiseur.validated_data))

    affectation, _ = AffectationProfilProjet.objects.update_or_create(
        projet_id=projet_id,
        profil_id=profil_id,
        clientele=serialiseur.validated_data["clientele"],
        defaults={
            "mode_facturation": request.data.get("mode_facturation", "journalier"),
            "charge_previsionnelle_jours": request.data.get("charge_previsionnelle_jours", 0) or 0,
            "coefficient_k": simulation["coefficients"]["coefficient_k_global"],
            "taux_horaire_recommande": simulation["resultats"]["taux_horaire_entreprise"],
            "taux_journalier_recommande": simulation["resultats"]["taux_journalier_entreprise"],
            "dernier_calcul": simulation,
            "observations": request.data.get("observations", ""),
        },
    )
    return Response(AffectationProfilProjetSerialiseur(affectation).data, status=status.HTTP_201_CREATED)


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
def vue_simuler_plan_activite(request):
    serialiseur = PlanActiviteEntreeSerialiseur(data=request.data)
    serialiseur.is_valid(raise_exception=True)

    lignes_simulation = []
    for ligne in serialiseur.validated_data["lignes"]:
        profil_code = ligne.get("profil_code")
        if profil_code:
            try:
                profil = ProfilMainOeuvre.objects.get(code=profil_code)
            except ProfilMainOeuvre.DoesNotExist:
                return Response(
                    {"detail": f"Profil inconnu : {profil_code}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            lignes_simulation.append(donnees_simulation_depuis_profil(profil, ligne) | {"effectif": ligne.get("effectif", 1)})
        else:
            lignes_simulation.append(ligne)

    return Response(simuler_plan_activite({"lignes": lignes_simulation}))


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
def vue_decomposer_prix_inverse(request):
    """
    Décompose intelligemment un prix de vente connu.
    Body: {prix_vente_unitaire, lot_type, methode, taux_personnalises?}
    """
    from .services import calculer_decomposition_inverse

    pv = request.data.get("prix_vente_unitaire")
    if not pv:
        return Response({"erreur": "prix_vente_unitaire est requis."}, status=400)

    try:
        resultat = calculer_decomposition_inverse(
            prix_vente_unitaire=pv,
            lot_type=request.data.get("lot_type", ""),
            methode=request.data.get("methode", "ratios_artiprix"),
            taux_personnalises=request.data.get("taux_personnalises"),
        )
        return Response(resultat)
    except Exception as exc:
        return Response({"erreur": str(exc)}, status=400)


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
def vue_estimation_tce(request):
    """
    Calcule une estimation TCE (Tous Corps d'État) par ratio depuis un montant global.

    Body: {
      type_ouvrage: str,       # batiment_collectif | batiment_tertiaire | erp | vrd
      montant_ht: float,       # Montant de base HT
      ajustements: dict,       # {code: ratio_pct} — surcharges optionnelles
    }

    Retourne la liste des corps d'état avec ratio % et montant HT.
    """
    RATIOS_TCE = {
        "batiment_collectif": {
            "label": "Bâtiment collectif résidentiel",
            "lignes": [
                {"code": "TCE-01", "designation": "VRD / Aménagements extérieurs", "unite": "fft", "ratio": 8.0, "categorie": "vrd"},
                {"code": "TCE-02", "designation": "Terrassements / Fondations", "unite": "fft", "ratio": 5.0, "categorie": "structure"},
                {"code": "TCE-03", "designation": "Gros œuvre / Structure béton armé", "unite": "fft", "ratio": 32.0, "categorie": "structure"},
                {"code": "TCE-04", "designation": "Charpente / Couverture", "unite": "fft", "ratio": 6.0, "categorie": "enveloppe"},
                {"code": "TCE-05", "designation": "Étanchéité", "unite": "fft", "ratio": 3.0, "categorie": "enveloppe"},
                {"code": "TCE-06", "designation": "Menuiseries extérieures / Façade", "unite": "fft", "ratio": 9.0, "categorie": "enveloppe"},
                {"code": "TCE-07", "designation": "Menuiseries intérieures / Serrurerie", "unite": "fft", "ratio": 4.0, "categorie": "second_oeuvre"},
                {"code": "TCE-08", "designation": "Isolation / Cloisons / Plâtrerie", "unite": "fft", "ratio": 10.0, "categorie": "second_oeuvre"},
                {"code": "TCE-09", "designation": "Revêtements de sol / Carrelage", "unite": "fft", "ratio": 4.0, "categorie": "second_oeuvre"},
                {"code": "TCE-10", "designation": "Peinture", "unite": "fft", "ratio": 3.0, "categorie": "second_oeuvre"},
                {"code": "TCE-11", "designation": "Électricité / Courants forts et faibles", "unite": "fft", "ratio": 7.0, "categorie": "fluides"},
                {"code": "TCE-12", "designation": "Plomberie / Sanitaires", "unite": "fft", "ratio": 5.0, "categorie": "fluides"},
                {"code": "TCE-13", "designation": "CVC / VMC", "unite": "fft", "ratio": 4.0, "categorie": "fluides"},
            ],
        },
        "batiment_tertiaire": {
            "label": "Bâtiment tertiaire / Bureaux",
            "lignes": [
                {"code": "TCE-01", "designation": "VRD / Aménagements extérieurs", "unite": "fft", "ratio": 6.0, "categorie": "vrd"},
                {"code": "TCE-02", "designation": "Terrassements / Fondations", "unite": "fft", "ratio": 5.0, "categorie": "structure"},
                {"code": "TCE-03", "designation": "Gros œuvre / Structure béton armé", "unite": "fft", "ratio": 28.0, "categorie": "structure"},
                {"code": "TCE-04", "designation": "Charpente / Couverture / Étanchéité", "unite": "fft", "ratio": 8.0, "categorie": "enveloppe"},
                {"code": "TCE-05", "designation": "Façades / Menuiseries extérieures", "unite": "fft", "ratio": 12.0, "categorie": "enveloppe"},
                {"code": "TCE-06", "designation": "Menuiseries intérieures / Cloisons", "unite": "fft", "ratio": 6.0, "categorie": "second_oeuvre"},
                {"code": "TCE-07", "designation": "Isolation / Faux plafonds", "unite": "fft", "ratio": 8.0, "categorie": "second_oeuvre"},
                {"code": "TCE-08", "designation": "Revêtements de sol", "unite": "fft", "ratio": 3.0, "categorie": "second_oeuvre"},
                {"code": "TCE-09", "designation": "Peinture", "unite": "fft", "ratio": 3.0, "categorie": "second_oeuvre"},
                {"code": "TCE-10", "designation": "Électricité / Courants forts et faibles", "unite": "fft", "ratio": 10.0, "categorie": "fluides"},
                {"code": "TCE-11", "designation": "Plomberie / Sanitaires", "unite": "fft", "ratio": 4.0, "categorie": "fluides"},
                {"code": "TCE-12", "designation": "CVC / GTB", "unite": "fft", "ratio": 7.0, "categorie": "fluides"},
            ],
        },
        "erp": {
            "label": "Équipement public / ERP",
            "lignes": [
                {"code": "TCE-01", "designation": "VRD / Aménagements extérieurs", "unite": "fft", "ratio": 10.0, "categorie": "vrd"},
                {"code": "TCE-02", "designation": "Terrassements / Fondations", "unite": "fft", "ratio": 7.0, "categorie": "structure"},
                {"code": "TCE-03", "designation": "Gros œuvre / Structure", "unite": "fft", "ratio": 28.0, "categorie": "structure"},
                {"code": "TCE-04", "designation": "Charpente bois / Couverture", "unite": "fft", "ratio": 10.0, "categorie": "enveloppe"},
                {"code": "TCE-05", "designation": "Étanchéité", "unite": "fft", "ratio": 2.0, "categorie": "enveloppe"},
                {"code": "TCE-06", "designation": "Menuiseries extérieures", "unite": "fft", "ratio": 8.0, "categorie": "enveloppe"},
                {"code": "TCE-07", "designation": "Menuiseries intérieures", "unite": "fft", "ratio": 4.0, "categorie": "second_oeuvre"},
                {"code": "TCE-08", "designation": "Isolation / Plâtrerie", "unite": "fft", "ratio": 9.0, "categorie": "second_oeuvre"},
                {"code": "TCE-09", "designation": "Revêtements de sol", "unite": "fft", "ratio": 3.0, "categorie": "second_oeuvre"},
                {"code": "TCE-10", "designation": "Peinture", "unite": "fft", "ratio": 3.0, "categorie": "second_oeuvre"},
                {"code": "TCE-11", "designation": "Électricité", "unite": "fft", "ratio": 8.0, "categorie": "fluides"},
                {"code": "TCE-12", "designation": "Plomberie / Sanitaires", "unite": "fft", "ratio": 4.0, "categorie": "fluides"},
                {"code": "TCE-13", "designation": "CVC / VMC", "unite": "fft", "ratio": 4.0, "categorie": "fluides"},
            ],
        },
        "vrd": {
            "label": "Travaux de VRD / Aménagement",
            "lignes": [
                {"code": "TCE-01", "designation": "Démolition / Démantèlement", "unite": "fft", "ratio": 5.0, "categorie": "vrd"},
                {"code": "TCE-02", "designation": "Terrassements généraux", "unite": "fft", "ratio": 20.0, "categorie": "vrd"},
                {"code": "TCE-03", "designation": "Fondations / Soutènements", "unite": "fft", "ratio": 8.0, "categorie": "structure"},
                {"code": "TCE-04", "designation": "Chaussées / Revêtements de voirie", "unite": "fft", "ratio": 25.0, "categorie": "vrd"},
                {"code": "TCE-05", "designation": "Réseaux eaux usées / pluviales", "unite": "fft", "ratio": 15.0, "categorie": "reseaux"},
                {"code": "TCE-06", "designation": "Réseau AEP / eau potable", "unite": "fft", "ratio": 8.0, "categorie": "reseaux"},
                {"code": "TCE-07", "designation": "Électricité / Éclairage public", "unite": "fft", "ratio": 10.0, "categorie": "reseaux"},
                {"code": "TCE-08", "designation": "Espaces verts / Plantations", "unite": "fft", "ratio": 5.0, "categorie": "paysager"},
                {"code": "TCE-09", "designation": "Mobilier urbain / Signalisation", "unite": "fft", "ratio": 4.0, "categorie": "paysager"},
            ],
        },
    }

    type_ouvrage = request.data.get("type_ouvrage", "batiment_collectif")
    if type_ouvrage not in RATIOS_TCE:
        type_ouvrage = "batiment_collectif"

    montant_ht = float(request.data.get("montant_ht", 0) or 0)
    if montant_ht <= 0:
        return Response({"erreur": "montant_ht doit être positif."}, status=400)

    ajustements = request.data.get("ajustements", {}) or {}
    referentiel = RATIOS_TCE[type_ouvrage]
    lignes_base = referentiel["lignes"]

    # Appliquer ajustements
    lignes = []
    for l in lignes_base:
        ratio = float(ajustements.get(l["code"], l["ratio"]))
        lignes.append({**l, "ratio": ratio})

    total_ratio = sum(l["ratio"] for l in lignes)

    COULEURS_CATEGORIE = {
        "vrd": "#6366f1",
        "structure": "#ef4444",
        "enveloppe": "#f59e0b",
        "second_oeuvre": "#10b981",
        "fluides": "#3b82f6",
        "reseaux": "#8b5cf6",
        "paysager": "#22c55e",
    }

    lignes_calculees = []
    for l in lignes:
        montant = montant_ht * (l["ratio"] / total_ratio)
        lignes_calculees.append({
            "code": l["code"],
            "designation": l["designation"],
            "unite": l["unite"],
            "categorie": l["categorie"],
            "ratio": round(l["ratio"], 2),
            "montant": round(montant, 2),
            "couleur": COULEURS_CATEGORIE.get(l["categorie"], "#64748b"),
        })

    return Response({
        "type_ouvrage": type_ouvrage,
        "label_ouvrage": referentiel["label"],
        "montant_ht": montant_ht,
        "total_ratio": round(total_ratio, 2),
        "lignes": lignes_calculees,
    })


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
def vue_creer_etude_tce(request):
    """
    Crée une EtudeEconomique avec des LignePrix générées par TCE.

    Body: {
      projet_id: str,
      intitule: str,
      type_ouvrage: str,
      montant_ht: float,
      ajustements: dict,  # optionnel
    }
    """
    from applications.projets.models import Projet

    projet_id = request.data.get("projet_id")
    if not projet_id:
        return Response({"erreur": "projet_id est requis."}, status=400)

    try:
        projet = Projet.objects.get(pk=projet_id)
    except Projet.DoesNotExist:
        return Response({"erreur": "Projet introuvable."}, status=404)

    # Réutiliser la logique de calcul TCE
    simulation_request = type("Req", (), {
        "data": {
            "type_ouvrage": request.data.get("type_ouvrage", "batiment_collectif"),
            "montant_ht": request.data.get("montant_ht"),
            "ajustements": request.data.get("ajustements", {}),
        }
    })()
    reponse_tce = vue_estimation_tce(simulation_request)
    if reponse_tce.status_code != 200:
        return reponse_tce
    tce = reponse_tce.data

    intitule = request.data.get("intitule") or f"Estimation TCE — {tce['label_ouvrage']}"

    etude = EtudeEconomique.objects.create(
        projet=projet,
        intitule=intitule,
        statut="brouillon",
        cree_par=request.user,
    )

    for i, ligne in enumerate(tce["lignes"], start=1):
        LignePrix.objects.create(
            etude=etude,
            numero_ordre=i,
            code=ligne["code"],
            designation=ligne["designation"],
            unite=ligne["unite"],
            quantite_prevue=Decimal("1"),
            cout_matieres=Decimal(str(round(ligne["montant"], 2))),
            debourse_sec_unitaire=Decimal(str(round(ligne["montant"], 2))),
            prix_vente_unitaire=Decimal(str(round(ligne["montant"], 2))),
            etat_rentabilite="indefini",
        )

    # Mettre à jour les totaux
    total = Decimal(str(tce["montant_ht"]))
    etude.total_debourse_sec = total
    etude.total_cout_direct = total
    etude.total_cout_revient = total
    etude.total_prix_vente = total
    etude.save()

    return Response({"id": str(etude.id), "intitule": etude.intitule}, status=201)


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated])
def vue_missions_par_type_client(request):
    """Retourne les missions économiques disponibles par type de client."""
    type_client = request.query_params.get("type_client", "")
    missions_map = {
        "moa_public": [
            {"code": "estimation_tce", "libelle": "Estimation TCE", "description": "Estimation toutes corps d'état en avant-projet"},
            {"code": "dpgf_lots", "libelle": "DPGF par lot", "description": "Décomposition du Prix Global et Forfaitaire par lot"},
            {"code": "dqe_consultation", "libelle": "DQE de consultation", "description": "Détail Quantitatif Estimatif pour la consultation"},
            {"code": "analyse_offres", "libelle": "Analyse d'offres", "description": "Analyse comparative des offres reçues"},
            {"code": "suivi_dgd", "libelle": "Suivi DGD", "description": "Décompte Général Définitif et liquidation"},
            {"code": "revision_prix", "libelle": "Révision de prix", "description": "Calcul de la révision de prix par formule paramétrique"},
            {"code": "controle_situation", "libelle": "Contrôle situations", "description": "Contrôle et visa des situations de travaux"},
        ],
        "moe": [
            {"code": "honoraires_moe", "libelle": "Honoraires MOE", "description": "Calcul des honoraires selon la mission et le montant travaux"},
            {"code": "planning_mission", "libelle": "Planning de mission", "description": "Jalons, livrables et délais de la mission"},
            {"code": "rapport_avancement", "libelle": "Rapport d'avancement", "description": "Synthèse mensuelle de l'avancement"},
            {"code": "mission_opc", "libelle": "Mission OPC", "description": "Ordonnancement, Pilotage et Coordination du chantier"},
        ],
        "entreprise_btp": [
            {"code": "etude_prix_analytique", "libelle": "Étude de prix analytique", "description": "Déboursé sec → Prix de vente par ressource"},
            {"code": "reponse_ao", "libelle": "Réponse AO", "description": "BPU, DQE et mémoire technique pour réponse à AO"},
            {"code": "marge_lot", "libelle": "Analyse de marge", "description": "Marges par ligne, lot et affaire"},
            {"code": "situation_travaux", "libelle": "Situations de travaux", "description": "Établissement et suivi des situations mensuelles"},
        ],
    }
    if type_client and type_client in missions_map:
        return Response(missions_map[type_client])
    return Response(missions_map)
