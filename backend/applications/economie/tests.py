from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from docx import Document as DocumentWord
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APIClient

from applications.batiment.models import LocalProgramme, ProgrammeBatiment
from applications.bibliotheque.models import LignePrixBibliotheque, SousDetailPrix
from applications.economie.models import (
    AffectationProfilProjet,
    AchatEtudePrix,
    ConventionCollective,
    EtudePrix,
    LignePrixEtude,
    ProfilMainOeuvre,
    ReferenceSocialeLocalisation,
    RegleConventionnelleProfil,
    VarianteLocaleRegleConventionnelle,
)
from applications.metres.models import LigneMetre, Metre
from applications.organisations.models import Organisation
from applications.projets.models import Projet


class WorkflowEtudePrixTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        utilisateur_model = get_user_model()

        self.organisation = Organisation.objects.create(
            code="ORG-TEST-ECO",
            nom="Organisation test économie",
            type_organisation="bureau_etudes",
            est_active=True,
        )
        self.utilisateur = utilisateur_model.objects.create_user(
            courriel="eco-tests@example.com",
            password="secret-test-123",
            prenom="Eco",
            nom="Tests",
            organisation=self.organisation,
            est_staff=True,
        )
        self.client.force_authenticate(self.utilisateur)

        self.projet = Projet.objects.create(
            reference="",
            intitule="Projet étude de prix",
            type_projet="autre",
            type_projet_autre="Audit études de prix",
            statut="en_cours",
            phase_actuelle="ao",
            organisation=self.organisation,
            maitre_ouvrage=self.organisation,
            responsable=self.utilisateur,
            cree_par=self.utilisateur,
            clientele_cible="entreprise_travaux",
            objectif_mission="devis_entreprise",
            commune="Mamoudzou",
            departement="976",
            montant_estime="98000.00",
        )
        self.projet_reference = Projet.objects.create(
            reference="REF-HIST-001",
            intitule="Projet historique comparable",
            type_projet="autre",
            type_projet_autre="Audit études de prix",
            statut="en_cours",
            phase_actuelle="ao",
            organisation=self.organisation,
            maitre_ouvrage=self.organisation,
            responsable=self.utilisateur,
            cree_par=self.utilisateur,
            clientele_cible="entreprise_travaux",
            objectif_mission="devis_entreprise",
            commune="Mamoudzou",
            departement="976",
            montant_marche="102000.00",
        )
        self.profil_maconnage = ProfilMainOeuvre.objects.create(
            code="MACON_TEST_API",
            libelle="Compagnon maçon test",
            categorie="compagnon",
            secteur_activite="batiment",
            metier="Maçon",
            niveau_classification="CP2",
            fonction_equipe="Compagnon",
            salaire_brut_mensuel_defaut="2350.00",
            taux_charges_patronales="0.4000",
            cout_equipement_mensuel="95.00",
            cout_transport_mensuel="80.00",
        )
        EtudePrix.objects.create(
            intitule="Étude historique lot VRD",
            code="EP-HIST-001",
            methode="analytique",
            lot_type="7.1",
            projet=self.projet_reference,
            organisation=self.organisation,
            zone_taux_horaire="A",
            taux_horaire_mo="41.0000",
            statut="publiee",
            prix_vente_ht="101500.0000",
            debourse_sec_ht="82000.0000",
        )
        self.programme = ProgrammeBatiment.objects.create(
            projet=self.projet,
            intitule="Programme enseignement",
            type_operation="construction_neuve",
            type_batiment="enseignement",
            shon_totale="1200.00",
        )
        LocalProgramme.objects.create(
            programme=self.programme,
            designation="Salle de classe",
            categorie="enseignement",
            nombre=6,
            surface_unitaire_m2="55.00",
        )
        self.programme_reference = ProgrammeBatiment.objects.create(
            projet=self.projet_reference,
            intitule="Programme enseignement comparable",
            type_operation="construction_neuve",
            type_batiment="enseignement",
            shon_totale="1180.00",
        )
        LocalProgramme.objects.create(
            programme=self.programme_reference,
            designation="Salle de classe",
            categorie="enseignement",
            nombre=5,
            surface_unitaire_m2="58.00",
        )
        self.metre = Metre.objects.create(
            projet=self.projet,
            intitule="Avant-métré façade",
            cree_par=self.utilisateur,
        )
        LigneMetre.objects.create(
            metre=self.metre,
            numero_ordre=1,
            designation="Enduit monocouche",
            nature="travaux",
            quantite="240.000",
            unite="m2",
            prix_unitaire_ht="32.5000",
        )
        self.metre_reference = Metre.objects.create(
            projet=self.projet_reference,
            intitule="Avant-métré façade référence",
            cree_par=self.utilisateur,
        )
        LigneMetre.objects.create(
            metre=self.metre_reference,
            numero_ordre=1,
            designation="Enduit monocouche",
            nature="travaux",
            quantite="230.000",
            unite="m2",
            prix_unitaire_ht="33.0000",
        )

    def creer_etude_prix(self, statut="brouillon"):
        reponse = self.client.post(
            "/api/economie/etudes-de-prix/",
            {
                "intitule": "Sous-détail analytique VRD",
                "code": "EP-TEST-001",
                "methode": "analytique",
                "lot_type": "7.1",
                "projet": str(self.projet.id),
                "organisation": str(self.organisation.id),
                "zone_taux_horaire": "A",
                "taux_horaire_mo": "41.0000",
                "statut": statut,
            },
            format="json",
        )
        self.assertEqual(reponse.status_code, status.HTTP_201_CREATED, reponse.data)
        return EtudePrix.objects.get(pk=reponse.data["id"])

    def ajouter_ligne_mo(self, etude: EtudePrix):
        reponse = self.client.post(
            f"/api/economie/etudes-de-prix/{etude.id}/lignes/",
            {
                "ordre": 1,
                "type_ressource": "mo",
                "designation": "Chef de chantier",
                "unite": "h",
                "quantite": "8",
                "taux_horaire": "45",
            },
            format="json",
        )
        self.assertEqual(reponse.status_code, status.HTTP_201_CREATED, reponse.data)
        return reponse.data["id"]

    def test_ajout_ligne_passe_letude_en_cours(self):
        etude = self.creer_etude_prix(statut="brouillon")

        self.ajouter_ligne_mo(etude)
        etude.refresh_from_db()

        self.assertEqual(etude.statut, "en_cours")
        self.assertEqual(str(etude.debourse_sec_ht), "360.0000")

    def test_validation_puis_publication_creent_une_entree_bibliotheque(self):
        etude = self.creer_etude_prix(statut="brouillon")
        self.ajouter_ligne_mo(etude)

        reponse_validation = self.client.post(
            f"/api/economie/etudes-de-prix/{etude.id}/valider/",
            {},
            format="json",
        )
        self.assertEqual(reponse_validation.status_code, status.HTTP_200_OK, reponse_validation.data)

        reponse_publication = self.client.post(
            f"/api/economie/etudes-de-prix/{etude.id}/publier/",
            {
                "unite": "m2",
                "quantite_ouvrage": "20",
            },
            format="json",
        )
        self.assertEqual(reponse_publication.status_code, status.HTTP_200_OK, reponse_publication.data)

        etude.refresh_from_db()
        self.assertEqual(etude.statut, "publiee")
        self.assertIsNotNone(etude.ligne_bibliotheque)

        ligne_bibliotheque = LignePrixBibliotheque.objects.get(pk=etude.ligne_bibliotheque_id)
        self.assertEqual(ligne_bibliotheque.code, etude.code)
        self.assertEqual(ligne_bibliotheque.unite, "m2")
        self.assertEqual(str(ligne_bibliotheque.temps_main_oeuvre), "0.4000")
        self.assertEqual(str(ligne_bibliotheque.debourse_sec_unitaire), "18.0000")
        self.assertEqual(SousDetailPrix.objects.filter(ligne_prix=ligne_bibliotheque).count(), 1)

        from applications.pieces_ecrites.models import ArticleCCTP

        article = ArticleCCTP.objects.get(ligne_prix_reference=ligne_bibliotheque)
        self.assertEqual(article.code_reference[:8], "LBH-CCTP")

    def test_ligne_mo_avec_profil_calcule_quantite_et_taux_recommande(self):
        etude = self.creer_etude_prix(statut="brouillon")

        reponse = self.client.post(
            f"/api/economie/etudes-de-prix/{etude.id}/lignes/",
            {
                "ordre": 1,
                "type_ressource": "mo",
                "profil_main_oeuvre": str(self.profil_maconnage.id),
                "designation": "Compagnon maçon analytique",
                "nombre_ressources": "2",
                "temps_unitaire": "1.5",
            },
            format="json",
        )

        self.assertEqual(reponse.status_code, status.HTTP_201_CREATED, reponse.data)
        ligne = LignePrixEtude.objects.get(pk=reponse.data["id"])
        self.assertEqual(str(ligne.quantite), "3.000000")
        self.assertEqual(str(ligne.nombre_ressources), "2.000")
        self.assertEqual(str(ligne.temps_unitaire), "1.500000")
        self.assertGreater(ligne.taux_horaire, 0)
        self.assertEqual(ligne.cout_unitaire_ht, ligne.taux_horaire)

    def test_une_etude_finalisee_ne_peut_plus_etre_modifiee(self):
        etude = self.creer_etude_prix(statut="brouillon")
        ligne_id = self.ajouter_ligne_mo(etude)

        reponse_validation = self.client.post(
            f"/api/economie/etudes-de-prix/{etude.id}/valider/",
            {},
            format="json",
        )
        self.assertEqual(reponse_validation.status_code, status.HTTP_200_OK, reponse_validation.data)

        reponse_modification_ligne = self.client.patch(
            f"/api/economie/etudes-de-prix/{etude.id}/lignes/{ligne_id}/",
            {"designation": "Modification interdite"},
            format="json",
        )
        self.assertEqual(reponse_modification_ligne.status_code, status.HTTP_400_BAD_REQUEST)

        reponse_modification_etude = self.client.patch(
            f"/api/economie/etudes-de-prix/{etude.id}/",
            {"intitule": "Modification interdite"},
            format="json",
        )
        self.assertEqual(reponse_modification_etude.status_code, status.HTTP_400_BAD_REQUEST)

    def test_export_xlsx_genere_un_bordereau_exploitable(self):
        etude = self.creer_etude_prix(statut="brouillon")
        self.ajouter_ligne_mo(etude)

        reponse = self.client.get(f"/api/economie/etudes-de-prix/{etude.id}/export/xlsx/")

        self.assertEqual(reponse.status_code, status.HTTP_200_OK)
        self.assertEqual(
            reponse["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        contenu = b"".join(reponse.streaming_content)
        classeur = load_workbook(filename=BytesIO(contenu), data_only=False)

        self.assertIn("Bordereau", classeur.sheetnames)
        self.assertIn("Synthèse", classeur.sheetnames)
        feuille = classeur["Bordereau"]
        self.assertEqual(feuille["B2"].value, "Chef de chantier")
        self.assertEqual(feuille["G2"].value, "=E2*F2")

    def test_recalcul_etude_prix_calcule_les_indicateurs_commerciaux(self):
        etude = self.creer_etude_prix(statut="brouillon")
        self.ajouter_ligne_mo(etude)
        etude.refresh_from_db()

        self.assertEqual(str(etude.debourse_sec_ht), "360.0000")
        self.assertEqual(str(etude.montant_frais_chantier_ht), "28.8000")
        self.assertEqual(str(etude.montant_frais_generaux_ht), "46.6560")
        self.assertEqual(str(etude.montant_aleas_ht), "10.8000")
        self.assertEqual(str(etude.cout_revient_ht), "446.2560")
        self.assertEqual(str(etude.marge_previsionnelle_ht), "44.6256")
        self.assertEqual(str(etude.prix_vente_ht), "490.8816")
        self.assertEqual(str(etude.coefficient_k), "1.3636")

    def test_proposition_achats_cree_des_lignes_conditionnees(self):
        etude = self.creer_etude_prix(statut="brouillon")
        reponse_ligne = self.client.post(
            f"/api/economie/etudes-de-prix/{etude.id}/lignes/",
            {
                "ordre": 1,
                "type_ressource": "matiere",
                "designation": "Béton C25/30",
                "unite": "m3",
                "quantite": "3.500000",
                "cout_unitaire_ht": "120.000000",
            },
            format="json",
        )
        self.assertEqual(reponse_ligne.status_code, status.HTTP_201_CREATED, reponse_ligne.data)

        reponse = self.client.post(
            f"/api/economie/etudes-de-prix/{etude.id}/achats/proposer/",
            {"remplacer": True},
            format="json",
        )

        self.assertEqual(reponse.status_code, status.HTTP_200_OK, reponse.data)
        achat = AchatEtudePrix.objects.get(etude=etude)
        self.assertEqual(achat.designation, "Béton C25/30")
        self.assertEqual(str(achat.quantite_besoin), "3.500000")
        self.assertEqual(str(achat.quantite_conditionnement), "3.500000")
        self.assertEqual(str(achat.nombre_conditionnements), "1.000")
        self.assertEqual(str(achat.cout_total_achat_ht), "120.0000")

    def test_cadrage_etude_prix_retourne_le_cadrage_metier(self):
        etude = self.creer_etude_prix(statut="brouillon")
        self.ajouter_ligne_mo(etude)

        reponse = self.client.get(f"/api/economie/etudes-de-prix/{etude.id}/cadrage/")

        self.assertEqual(reponse.status_code, status.HTTP_200_OK, reponse.data)
        self.assertTrue(reponse.data["questions"])
        self.assertTrue(reponse.data["methodes_recommandees"])
        self.assertEqual(reponse.data["sous_detail_disponible"], True)

    def test_comparatif_etude_prix_retourne_ratio_rex_et_projets_similaires(self):
        etude = self.creer_etude_prix(statut="brouillon")
        self.ajouter_ligne_mo(etude)

        reponse = self.client.get(f"/api/economie/etudes-de-prix/{etude.id}/comparatif/")

        self.assertEqual(reponse.status_code, status.HTTP_200_OK, reponse.data)
        self.assertTrue(reponse.data["projets_similaires"])
        self.assertGreater(float(reponse.data["taux_similarite_moyen"]), 0)
        self.assertGreater(float(reponse.data["estimation_ratio_ht"]), 0)
        self.assertGreater(float(reponse.data["estimation_rex_ht"]), 0)

    def test_comparatif_etude_prix_integre_metres_et_programme_batiment(self):
        etude = self.creer_etude_prix(statut="brouillon")
        self.ajouter_ligne_mo(etude)

        reponse = self.client.get(f"/api/economie/etudes-de-prix/{etude.id}/comparatif/")

        self.assertEqual(reponse.status_code, status.HTTP_200_OK, reponse.data)
        self.assertEqual(reponse.data["profil_projet"]["types_batiment"], ["enseignement"])
        self.assertEqual(reponse.data["profil_projet"]["types_operation_batiment"], ["construction_neuve"])
        self.assertEqual(reponse.data["profil_projet"]["categories_locaux"], ["enseignement"])
        self.assertEqual(float(reponse.data["profil_projet"]["montant_metres_ht"]), 7800.0)
        self.assertEqual(reponse.data["profil_projet"]["nombre_lignes_metre"], 1)
        self.assertGreater(float(reponse.data["taux_similarite_moyen"]), 0)

    def test_export_comparatif_xlsx_genere_un_graphique_et_des_syntheses(self):
        etude = self.creer_etude_prix(statut="brouillon")
        self.ajouter_ligne_mo(etude)

        reponse = self.client.get(f"/api/economie/etudes-de-prix/{etude.id}/export/comparatif-xlsx/")

        self.assertEqual(reponse.status_code, status.HTTP_200_OK)
        contenu = b"".join(reponse.streaming_content)
        classeur = load_workbook(filename=BytesIO(contenu), data_only=False)
        self.assertIn("Comparatif", classeur.sheetnames)
        self.assertIn("Projets similaires", classeur.sheetnames)
        self.assertIn("Alertes", classeur.sheetnames)

    def test_export_achats_xlsx_genere_un_bon_de_commande(self):
        etude = self.creer_etude_prix(statut="brouillon")
        AchatEtudePrix.objects.create(
            etude=etude,
            ordre=1,
            designation="Treillis soudé",
            fournisseur="Fournisseur test",
            unite_achat="panneau",
            quantite_besoin="3.000000",
            quantite_conditionnement="2.000000",
            prix_unitaire_achat_ht="45.000000",
        )

        reponse = self.client.get(f"/api/economie/etudes-de-prix/{etude.id}/export/achats-xlsx/")

        self.assertEqual(reponse.status_code, status.HTTP_200_OK)
        contenu = b"".join(reponse.streaming_content)
        classeur = load_workbook(filename=BytesIO(contenu), data_only=False)
        feuille = classeur["Bon de commande"]
        self.assertEqual(feuille["B2"].value, "Treillis soudé")

    def test_export_note_moa_docx_genere_un_livrable_oriente_moa(self):
        etude = self.creer_etude_prix(statut="brouillon")
        self.ajouter_ligne_mo(etude)

        reponse = self.client.get(f"/api/economie/etudes-de-prix/{etude.id}/export/note-moa-docx/")

        self.assertEqual(reponse.status_code, status.HTTP_200_OK)
        self.assertEqual(
            reponse["Content-Type"],
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )
        contenu = b"".join(reponse.streaming_content)
        document = DocumentWord(BytesIO(contenu))
        texte = "\n".join(paragraphe.text for paragraphe in document.paragraphs)
        self.assertIn("Note de vérification d'enveloppe", texte)
        self.assertIn("Positionnement maîtrise d'ouvrage", texte)
        self.assertIn("Recommandation MOA", texte)

    def test_export_note_moe_docx_genere_un_livrable_oriente_moe(self):
        etude = self.creer_etude_prix(statut="brouillon")
        self.ajouter_ligne_mo(etude)

        reponse = self.client.get(f"/api/economie/etudes-de-prix/{etude.id}/export/note-moe-docx/")

        self.assertEqual(reponse.status_code, status.HTTP_200_OK)
        self.assertEqual(
            reponse["Content-Type"],
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )
        contenu = b"".join(reponse.streaming_content)
        document = DocumentWord(BytesIO(contenu))
        texte = "\n".join(paragraphe.text for paragraphe in document.paragraphs)
        self.assertIn("Note d'estimation consolidée", texte)
        self.assertIn("Positionnement maîtrise d'œuvre", texte)
        self.assertIn("Recommandation MOE", texte)


class SimulationMainOeuvreTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        utilisateur_model = get_user_model()

        self.organisation = Organisation.objects.create(
            code="ORG-TEST-SIMU",
            nom="Organisation test simulateur",
            type_organisation="bureau_etudes",
            est_active=True,
        )
        self.utilisateur = utilisateur_model.objects.create_user(
            courriel="simu@example.com",
            password="secret-test-123",
            prenom="Simu",
            nom="Tests",
            organisation=self.organisation,
            est_staff=True,
        )
        self.client.force_authenticate(self.utilisateur)

        self.projet = Projet.objects.create(
            intitule="Projet simulateur",
            type_projet="autre",
            type_projet_autre="Simulation économique",
            statut="en_cours",
            phase_actuelle="ao",
            organisation=self.organisation,
            maitre_ouvrage=self.organisation,
            responsable=self.utilisateur,
            cree_par=self.utilisateur,
        )
        self.profil = ProfilMainOeuvre.objects.create(
            code="INGE_TEST",
            libelle="Ingénieur test",
            categorie="ingenieur",
            salaire_brut_mensuel_defaut="3600.00",
        )
        self.convention = ConventionCollective.objects.create(
            code="BET_TEST",
            libelle="Convention test bureau d'études",
            idcc="1486",
            localisation="mayotte",
        )
        self.regle = RegleConventionnelleProfil.objects.create(
            convention=self.convention,
            code="INGE_TEST_CADRE",
            libelle="Ingénieur cadre test",
            categorie="ingenieur",
            statut_cadre=True,
            salaire_brut_minimum_mensuel="3900.00",
            mutuelle_employeur_mensuelle_defaut="70.00",
            titres_restaurant_employeur_mensuels_defaut="95.00",
            prime_transport_mensuelle_defaut="40.00",
            taux_absenteisme_defaut="0.0200",
            taux_temps_improductif_defaut="0.1400",
            cout_recrutement_initial_defaut="3200.00",
        )
        self.profil.convention_collective = self.convention
        self.profil.regle_conventionnelle = self.regle
        self.profil.save(update_fields=["convention_collective", "regle_conventionnelle"])
        self.reference_mayotte, _ = ReferenceSocialeLocalisation.objects.update_or_create(
            localisation="mayotte",
            defaults={
                "code": "TEST_MAYOTTE",
                "libelle": "Référentiel Mayotte test",
                "smic_horaire": "9.33",
                "heures_legales_mensuelles": "151.67",
                "commentaire_reglementaire": "Référentiel Mayotte de test",
                "source_officielle": "https://entreprendre.service-public.fr/vosdroits/F31326",
                "est_active": True,
            },
        )
        self.variante_mayotte, _ = VarianteLocaleRegleConventionnelle.objects.update_or_create(
            regle=self.regle,
            localisation="mayotte",
            defaults={
                "libelle": "Ingénieur cadre Mayotte",
                "salaire_brut_minimum_mensuel": "3300.00",
                "taux_charges_patronales_defaut": "0.3500",
                "mutuelle_employeur_mensuelle_defaut": "45.00",
                "taux_occupation_facturable_defaut": "0.8200",
                "observations": "Variante Mayotte de test",
                "source_officielle": "https://entreprendre.service-public.fr/vosdroits/F31326",
                "est_active": True,
            },
        )

    def charge_simulation(self):
        return {
            "profil_code": self.profil.code,
            "profil_libelle": self.profil.libelle,
            "clientele": "public",
            "localisation": "mayotte",
            "contrat_travail": "cdi",
            "statut_cadre": True,
            "quotite_travail": "1.0000",
            "salaire_brut_mensuel": "3600.00",
            "primes_mensuelles": "250.00",
            "avantages_mensuels": "120.00",
            "heures_supplementaires_mensuelles": "4.00",
            "majoration_heures_supplementaires": "0.2500",
            "heures_contractuelles_mensuelles": "151.67",
            "heures_par_jour": "7.00",
            "taux_charges_salariales": "0.2200",
            "taux_charges_patronales": "0.4400",
            "taux_absenteisme": "0.0300",
            "taux_temps_improductif": "0.1200",
            "taux_frais_agence": "0.1300",
            "taux_risque_operationnel": "0.0200",
            "taux_marge_cible": "0.0800",
            "mutuelle_employeur_mensuelle": "55.00",
            "titres_restaurant_employeur_mensuels": "90.00",
            "prime_transport_mensuelle": "45.00",
            "cout_equipement_mensuel": "60.00",
            "cout_transport_mensuel": "80.00",
            "cout_structure_mensuel": "340.00",
            "appliquer_rgdu": True,
            "taux_occupation_facturable": "0.7600",
            "jours_facturables_cibles_annuels": "150.00",
            "cout_recrutement_initial": "2500.00",
        }

    def test_simulation_main_oeuvre_renvoie_k_et_taux(self):
        reponse = self.client.post(
            "/api/economie/simulateur-main-oeuvre/",
            self.charge_simulation(),
            format="json",
        )

        self.assertEqual(reponse.status_code, status.HTTP_200_OK, reponse.data)
        self.assertEqual(reponse.data["profil_libelle"], "Ingénieur test")
        self.assertGreater(reponse.data["coefficients"]["coefficient_k_global"], 1)
        self.assertIn("projection_annuelle", reponse.data)
        self.assertIn("hypotheses_reglementaires", reponse.data)
        self.assertEqual(reponse.data["convention_collective_libelle"], self.convention.libelle)
        self.assertEqual(reponse.data["regle_conventionnelle_libelle"], self.regle.libelle)
        self.assertGreater(
            reponse.data["resultats"]["taux_journalier_entreprise"],
            reponse.data["resultats"]["cout_journalier_productif"],
        )

    def test_simulation_signale_minimum_conventionnel_si_brut_insuffisant(self):
        charge = self.charge_simulation()
        charge["salaire_brut_mensuel"] = "3200.00"
        charge["mutuelle_employeur_mensuelle"] = "0.00"
        charge["titres_restaurant_employeur_mensuels"] = "0.00"
        charge["prime_transport_mensuelle"] = "0.00"
        charge["cout_recrutement_initial"] = "0.00"

        reponse = self.client.post(
            "/api/economie/simulateur-main-oeuvre/",
            charge,
            format="json",
        )

        self.assertEqual(reponse.status_code, status.HTTP_200_OK, reponse.data)
        self.assertEqual(
            reponse.data["hypotheses_reglementaires"]["minimum_conventionnel_reference"],
            3300.0,
        )
        self.assertEqual(reponse.data["bulletin"]["mutuelle_employeur_mensuelle"], 0.0)
        self.assertTrue(
            any("minimum conventionnel" in avertissement.lower() for avertissement in reponse.data["avertissements"])
        )

    def test_defauts_simulation_profil_appliquent_la_variante_locale_mayotte(self):
        reponse = self.client.get(
            f"/api/economie/profils-main-oeuvre/{self.profil.id}/simulation-defauts/?localisation=mayotte"
        )

        self.assertEqual(reponse.status_code, status.HTTP_200_OK, reponse.data)
        self.assertEqual(reponse.data["reference_sociale_localisation_libelle"], self.reference_mayotte.libelle)
        self.assertEqual(reponse.data["variante_locale_regle_libelle"], self.variante_mayotte.libelle)
        self.assertEqual(str(reponse.data["mutuelle_employeur_mensuelle"]), "45.00")
        self.assertEqual(str(reponse.data["taux_charges_patronales"]), "0.3500")
        self.assertEqual(str(reponse.data["salaire_brut_minimum_conventionnel"]), "3300.00")

    def test_export_pdf_simulation_main_oeuvre(self):
        reponse = self.client.post(
            "/api/economie/simulateur-main-oeuvre/export/pdf/",
            self.charge_simulation(),
            format="json",
        )

        self.assertEqual(reponse.status_code, status.HTTP_200_OK)
        self.assertEqual(reponse["Content-Type"], "application/pdf")

    def test_creation_affectation_depuis_simulation(self):
        charge = self.charge_simulation()
        charge.update(
            {
                "projet": str(self.projet.id),
                "profil": str(self.profil.id),
                "mode_facturation": "journalier",
                "charge_previsionnelle_jours": "18.00",
            }
        )

        reponse = self.client.post(
            "/api/economie/simulateur-main-oeuvre/affecter/",
            charge,
            format="json",
        )

        self.assertEqual(reponse.status_code, status.HTTP_201_CREATED, reponse.data)
        affectation = AffectationProfilProjet.objects.get(pk=reponse.data["id"])
        self.assertEqual(affectation.profil_id, self.profil.id)
        self.assertEqual(affectation.projet_id, self.projet.id)
        self.assertGreater(float(affectation.taux_journalier_recommande), 0)

    def test_simulation_plan_activite_multi_profils(self):
        reponse = self.client.post(
            "/api/economie/pilotage-activite/simuler/",
            {
                "lignes": [
                    {
                        **self.charge_simulation(),
                        "effectif": 2,
                    },
                    {
                        **self.charge_simulation(),
                        "profil_code": self.profil.code,
                        "profil_libelle": "Ingénieur test 2",
                        "salaire_brut_mensuel": "2900.00",
                        "effectif": 1,
                        "clientele": "particulier_pme",
                    },
                ]
            },
            format="json",
        )

        self.assertEqual(reponse.status_code, status.HTTP_200_OK, reponse.data)
        self.assertEqual(len(reponse.data["lignes"]), 2)
        self.assertGreater(reponse.data["totaux"]["chiffre_affaires_cible_annuel"], 0)
        self.assertGreater(reponse.data["totaux"]["cout_total_annuel"], 0)

    def test_liste_profils_expose_convention_et_regle(self):
        reponse = self.client.get("/api/economie/profils-main-oeuvre/?actifs=1")

        self.assertEqual(reponse.status_code, status.HTTP_200_OK, reponse.data)
        resultats = reponse.data["results"] if isinstance(reponse.data, dict) and "results" in reponse.data else reponse.data
        profil = next((item for item in resultats if item["code"] == self.profil.code), None)
        self.assertIsNotNone(profil)
        self.assertEqual(profil["convention_collective_libelle"], self.convention.libelle)
        self.assertEqual(profil["regle_conventionnelle_libelle"], self.regle.libelle)

    def test_liste_references_sociales_localisation(self):
        reponse = self.client.get("/api/economie/references-sociales-localisation/?actifs=1")

        self.assertEqual(reponse.status_code, status.HTTP_200_OK, reponse.data)
        resultats = reponse.data["results"] if isinstance(reponse.data, dict) and "results" in reponse.data else reponse.data
        reference = next((item for item in resultats if item["code"] == self.reference_mayotte.code), None)
        self.assertIsNotNone(reference)
        self.assertEqual(reference["localisation"], "mayotte")
