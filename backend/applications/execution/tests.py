from django.contrib.auth import get_user_model
from django.test import TestCase
from rest_framework import status
from rest_framework.test import APIClient
from openpyxl import load_workbook
from io import BytesIO

from applications.economie.models import EtudeEconomique, LignePrix, ProfilMainOeuvre
from applications.execution.models import PlanningChantier, SuiviExecution
from applications.organisations.models import Organisation
from applications.projets.models import Projet


class PlanningChantierTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        utilisateur_model = get_user_model()

        self.organisation = Organisation.objects.create(
            code="ORG-TEST-EXEC",
            nom="Organisation test exécution",
            type_organisation="bureau_etudes",
            est_active=True,
        )
        self.utilisateur = utilisateur_model.objects.create_user(
            courriel="execution@example.com",
            password="secret-test-123",
            prenom="Exec",
            nom="Tests",
            organisation=self.organisation,
            est_staff=True,
        )
        self.client.force_authenticate(self.utilisateur)

        self.projet = Projet.objects.create(
            intitule="Projet planning chantier",
            type_projet="travaux",
            statut="en_cours",
            phase_actuelle="exe",
            organisation=self.organisation,
            maitre_ouvrage=self.organisation,
            responsable=self.utilisateur,
            cree_par=self.utilisateur,
        )
        self.suivi = SuiviExecution.objects.create(
            projet=self.projet,
            entreprise_principale=self.organisation,
        )
        self.profil = ProfilMainOeuvre.objects.create(
            code="OUV_TEST",
            libelle="Ouvrier test",
            categorie="ouvrier",
        )
        self.etude = EtudeEconomique.objects.create(
            projet=self.projet,
            intitule="DPGF travaux de test",
            cree_par=self.utilisateur,
        )
        LignePrix.objects.create(
            etude=self.etude,
            numero_ordre=1,
            code="T1",
            designation="Terrassement général",
            unite="m3",
            quantite_prevue="10.000",
            temps_main_oeuvre="2.0000",
        )
        LignePrix.objects.create(
            etude=self.etude,
            numero_ordre=2,
            code="T2",
            designation="Béton de propreté",
            unite="m3",
            quantite_prevue="5.000",
            temps_main_oeuvre="1.0000",
        )

    def test_generation_planning_depuis_etude_economique(self):
        reponse = self.client.post(
            f"/api/execution/{self.suivi.id}/plannings/",
            {
                "intitule": "Planning DPGF chantier",
                "source_donnees": "etude_economique",
                "etude_economique": str(self.etude.id),
                "date_debut_reference": "2026-04-06",
                "heures_par_jour": "7.00",
                "coefficient_rendement_global": "1.0000",
            },
            format="json",
        )

        self.assertEqual(reponse.status_code, status.HTTP_201_CREATED, reponse.data)

        planning = PlanningChantier.objects.prefetch_related("taches").get(pk=reponse.data["id"])
        taches = list(planning.taches.order_by("numero_ordre"))

        self.assertEqual(len(taches), 2)
        self.assertEqual(taches[0].designation, "Terrassement général")
        self.assertEqual(str(taches[0].duree_jours), "2.86")
        self.assertEqual(str(taches[1].duree_jours), "1.00")
        self.assertEqual(len(planning.chemin_critique), 2)
        self.assertEqual(planning.synthese_calcul["nb_taches"], 2)
        self.assertEqual(planning.synthese_calcul["nb_taches_critiques"], 2)
        self.assertEqual(str(planning.synthese_calcul["duree_totale_jours"]), "3.86")
        self.assertEqual(str(taches[0].date_debut_calculee), "2026-04-06")
        self.assertEqual(str(taches[1].date_debut_calculee), "2026-04-09")

    def test_affectation_equipe_reduit_la_duree_de_tache(self):
        reponse = self.client.post(
            f"/api/execution/{self.suivi.id}/plannings/",
            {
                "intitule": "Planning DPGF chantier",
                "source_donnees": "etude_economique",
                "etude_economique": str(self.etude.id),
                "date_debut_reference": "2026-04-06",
            },
            format="json",
        )
        self.assertEqual(reponse.status_code, status.HTTP_201_CREATED, reponse.data)

        planning = PlanningChantier.objects.prefetch_related("taches").get(pk=reponse.data["id"])
        premiere_tache = planning.taches.order_by("numero_ordre").first()
        self.assertIsNotNone(premiere_tache)

        reponse_affectation = self.client.put(
            f"/api/execution/plannings/{planning.id}/taches/{premiere_tache.id}/affectations/",
            {
                "affectations": [
                    {
                        "profil": str(self.profil.id),
                        "effectif": 2,
                        "rendement_relatif": "1.0000",
                        "est_chef_equipe": True,
                    }
                ]
            },
            format="json",
        )
        self.assertEqual(reponse_affectation.status_code, status.HTTP_200_OK, reponse_affectation.data)

        premiere_tache.refresh_from_db()
        planning.refresh_from_db()

        self.assertEqual(premiere_tache.effectif_alloue, 2)
        self.assertEqual(str(premiere_tache.duree_jours), "1.43")
        self.assertEqual(str(planning.synthese_calcul["duree_totale_jours"]), "2.43")

    def test_calendrier_ouvre_reporte_les_dates_hors_weekend(self):
        reponse = self.client.post(
            f"/api/execution/{self.suivi.id}/plannings/",
            {
                "intitule": "Planning jours ouvrés",
                "source_donnees": "etude_economique",
                "etude_economique": str(self.etude.id),
                "date_debut_reference": "2026-04-03",
            },
            format="json",
        )
        self.assertEqual(reponse.status_code, status.HTTP_201_CREATED, reponse.data)

        planning = PlanningChantier.objects.prefetch_related("taches").get(pk=reponse.data["id"])
        premiere_tache, seconde_tache = planning.taches.order_by("numero_ordre")

        self.assertEqual(str(premiere_tache.date_debut_calculee), "2026-04-03")
        self.assertEqual(str(seconde_tache.date_debut_calculee), "2026-04-08")

    def test_dependance_editable_et_export_xlsx(self):
        reponse = self.client.post(
            f"/api/execution/{self.suivi.id}/plannings/",
            {
                "intitule": "Planning export",
                "source_donnees": "etude_economique",
                "etude_economique": str(self.etude.id),
                "date_debut_reference": "2026-04-06",
            },
            format="json",
        )
        self.assertEqual(reponse.status_code, status.HTTP_201_CREATED, reponse.data)

        planning = PlanningChantier.objects.prefetch_related("taches").get(pk=reponse.data["id"])
        premiere_tache, seconde_tache = planning.taches.order_by("numero_ordre")

        dependance_existante = seconde_tache.dependances_entrantes.first()
        self.assertIsNotNone(dependance_existante)
        reponse_dependance = self.client.patch(
            f"/api/execution/plannings/{planning.id}/dependances/{dependance_existante.id}/",
            {"type_dependance": "dd", "decalage_jours": "1.00"},
            format="json",
        )
        self.assertEqual(reponse_dependance.status_code, status.HTTP_200_OK, reponse_dependance.data)

        reponse_export = self.client.get(f"/api/execution/plannings/{planning.id}/export/xlsx/")
        self.assertEqual(reponse_export.status_code, status.HTTP_200_OK)
        self.assertEqual(
            reponse_export["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        contenu = b"".join(reponse_export.streaming_content)
        classeur = load_workbook(filename=BytesIO(contenu), data_only=False)
        self.assertIn("Planning", classeur.sheetnames)
        feuille = classeur["Planning"]
        self.assertEqual(feuille["C2"].value, "Terrassement général")
        self.assertEqual(feuille["M3"].value, "T1 Début -> Début (1.00 j)")

    def test_export_pdf_planning(self):
        reponse = self.client.post(
            f"/api/execution/{self.suivi.id}/plannings/",
            {
                "intitule": "Planning PDF",
                "source_donnees": "etude_economique",
                "etude_economique": str(self.etude.id),
                "date_debut_reference": "2026-04-06",
            },
            format="json",
        )
        self.assertEqual(reponse.status_code, status.HTTP_201_CREATED, reponse.data)

        planning_id = reponse.data["id"]
        reponse_pdf = self.client.get(f"/api/execution/plannings/{planning_id}/export/pdf/")

        self.assertEqual(reponse_pdf.status_code, status.HTTP_200_OK)
        self.assertEqual(reponse_pdf["Content-Type"], "application/pdf")
